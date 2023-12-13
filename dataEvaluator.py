import xlsxwriter
import matplotlib.pyplot as plt
import openpyxl as xlsx


class Evaluator:

    def __init__(self):
        self.allRoads = [
            "vehicle1_1_6_1",
            "vehicle1_2_6_2",
            "vehicle7_4",
            "vehicle5_2_2",
            "vehicle3_2_8",

            "vehicle1_1_8",
            "vehicle7_6_1",
            "vehicle5_1_4",
            "vehicle3_1_2",

            "vehicle1_2_4",
            "vehicle7_2",
            "vehicle3_2_6_2",
            "vehicle5_3_8"]
        self.allRoadsNames = {
            "vehicle1_1_6_1": "road 1-3l",
            "vehicle1_2_6_2": "road 1-3r",
            "vehicle7_4": "road 4-2",
            "vehicle5_2_2": "road 3-1",
            "vehicle3_2_8": "road 2-4",

            "vehicle1_1_8": "road 1-4",
            "vehicle7_6_1": "road 4-3",
            "vehicle5_1_4": "road 3-2",
            "vehicle3_1_2": "road 2-1",

            "vehicle1_2_4": "road 1-2",
            "vehicle7_2": "road 4-1",
            "vehicle3_2_6_2": "road 2-3",
            "vehicle5_3_8": "road 3-4"}
        self.allDrivesTime = {}
        self.resultDate = {}
        for road in self.allRoads:
            data1 = {"numberOfCars": 0, "generalDriveSpan": 0, "stopSpan": 0, "realStopSpan": 0, "realDriveSpan": 0}
            self.allDrivesTime[road] = data1
            data2 = {"numberOfCars": 0, "generalDriveSpan": 0, "stopSpan": 0, "realStopSpan": 0, "realDriveSpan": 0}
            self.resultDate[road] = data2

    def barDiagramDraw(self, path="output/test/path.png", title="test", xLabel="x", yLabel="y", data=[]):
        roads = self.allRoads.copy()
        for i in range(0, len(roads)):
            roads[i] = self.allRoadsNames[roads[i]]

        plt.bar(roads, data[2])
        plt.title(title)
        plt.xlabel(xLabel)
        plt.ylabel(yLabel)
        plt.xticks(range(len(roads)), roads, rotation=90)
        plt.subplots_adjust(bottom=0.25)
        if title == "numberOfCars":
            plt.yticks(range(0, 180, 10))

        def addLabels(x, y):
            for index in range(len(x)):
                plt.text(index, y[index], y[index])

        addLabels(roads, data[2])
        plt.savefig(path)
        plt.show()

    def dataEvaluate1(self, factor="realTime"):
        for n in range(1, 6):
            r = 0

            for r in range(0, 20):

                path = "output/" + str(factor) + "/" + str(factor) + str(n) + "/" + str(r) + "/statisticsResult.xlsx"

                wb = xlsx.load_workbook(path)
                sh = wb.active
                x = 11

                for road in self.allDrivesTime:
                    y = 2
                    cell = sh.cell(row=x, column=y)
                    self.allDrivesTime[road]["numberOfCars"] += cell.value
                    y += 1

                    cell = sh.cell(row=x, column=y)
                    self.allDrivesTime[road]["generalDriveSpan"] += cell.value
                    y += 1

                    cell = sh.cell(row=x, column=y)
                    self.allDrivesTime[road]["stopSpan"] += cell.value
                    y += 1

                    cell = sh.cell(row=x, column=y)
                    self.allDrivesTime[road]["realStopSpan"] += cell.value
                    y += 1

                    cell = sh.cell(row=x, column=y)
                    self.allDrivesTime[road]["realDriveSpan"] += cell.value

                    x += 1

                r += 1

            for road in self.allDrivesTime:
                for k in self.allDrivesTime[road]:
                    self.allDrivesTime[road][k] /= r

            path = "output/" + str(factor) + "/" + str(factor) + str(n) + "/general/general.xlsx"

            workbook = xlsxwriter.Workbook(path)
            worksheet = workbook.add_worksheet()

            worksheet.write(2, 1, "numberOfCars")
            worksheet.write(2, 2, "generalDriveSpan")
            worksheet.write(2, 3, "stopSpan")
            worksheet.write(2, 4, "realStopSpan")
            worksheet.write(2, 5, "realDriveSpan")

            x = 3
            numberOfCars = []
            generalDriveSpan = []
            stopSpan = []
            realStopSpan = []
            realDriveSpan = []

            for road in self.allDrivesTime:
                y = 0
                worksheet.write(x, y, road)
                numberOfCars.append(int(self.allDrivesTime[road]["numberOfCars"]))
                generalDriveSpan.append(round(self.allDrivesTime[road]["generalDriveSpan"], 2))
                stopSpan.append(round(self.allDrivesTime[road]["stopSpan"], 2))
                realStopSpan.append(round(self.allDrivesTime[road]["realStopSpan"], 2))
                realDriveSpan.append(round(self.allDrivesTime[road]["realDriveSpan"], 2))
                flag = True
                for k in self.allDrivesTime[road]:
                    y += 1
                    val = self.allDrivesTime[road][k]  # 29.12.2022
                    if flag:
                        val = round(val)
                        flag = False
                    worksheet.write(x, y, val)
                x += 1
            workbook.close()

            self.barDiagramDraw(path="output/" + str(factor) + "/" + str(factor) + str(n) + "/general/numberOfCars.png",
                                title="numberOfCars", xLabel="road", yLabel="numberOfCars",
                                data=["numberOfCars", self.allRoads, numberOfCars])
            self.barDiagramDraw(
                path="output/" + str(factor) + "/" + str(factor) + str(n) + "/general/generalDriveSpan.png",
                title="generalDriveSpan", xLabel="road", yLabel="average of generalDriveSpan in S ",
                data=["generalDriveSpan", self.allRoads, generalDriveSpan])
            self.barDiagramDraw(path="output/" + str(factor) + "/" + str(factor) + str(n) + "/general/stopSpan.png",
                                title="stopSpan", xLabel="road", yLabel="average of stopSpan in S",
                                data=["stopSpan", self.allRoads, stopSpan])
            self.barDiagramDraw(path="output/" + str(factor) + "/" + str(factor) + str(n) + "/general/realStopSpan.png",
                                title="realStopSpan", xLabel="road", yLabel="average of realStopSpan in S",
                                data=["realStopSpan", self.allRoads, realStopSpan])
            self.barDiagramDraw(
                path="output/" + str(factor) + "/" + str(factor) + str(n) + "/general/realDriveSpan.png",
                title="realDriveSpan", xLabel="road", yLabel="average of realDriveSpan in S",
                data=["realDriveSpan", self.allRoads, realDriveSpan])

    def dataEvaluate2(self, factor="realTime"):
        for i in range(1, 6):

            path = "output/" + str(factor) + "/" + str(factor)
            path += str(i) + "/" + "general/general.xlsx"
            wb = xlsx.load_workbook(path)
            sh = wb.active
            x = 4

            for road in self.resultDate:
                y = 2
                cell = sh.cell(row=x, column=y)
                self.resultDate[road]["numberOfCars"] += cell.value
                y += 1

                cell = sh.cell(row=x, column=y)
                self.resultDate[road]["generalDriveSpan"] += cell.value
                y += 1

                cell = sh.cell(row=x, column=y)
                self.resultDate[road]["stopSpan"] += cell.value
                y += 1

                cell = sh.cell(row=x, column=y)
                self.resultDate[road]["realStopSpan"] += cell.value
                y += 1

                cell = sh.cell(row=x, column=y)
                self.resultDate[road]["realDriveSpan"] += cell.value

                x += 1

        for road in self.resultDate:
            for k in self.resultDate[road]:
                self.resultDate[road][k] /= 5

        path = "output/" + str(factor) + "/general/" + str(factor) + "_general.xlsx"

        workbook = xlsxwriter.Workbook(path)
        worksheet = workbook.add_worksheet()

        # worksheet.write(2, 1, "numberOfCars")
        p = 1
        worksheet.write(2, p, "generalDriveSpan")
        worksheet.write(2, p + 1, "stopSpan")
        worksheet.write(2, p + 2, "realStopSpan")
        worksheet.write(2, p + 3, "realDriveSpan")

        x = 3
        numberOfCars = []
        generalDriveSpan = []
        stopSpan = []
        realStopSpan = []
        realDriveSpan = []

        for road in self.resultDate:
            y = 0
            worksheet.write(x, y, self.allRoadsNames[road])
            numberOfCars.append(int(self.resultDate[road]["numberOfCars"]))
            generalDriveSpan.append(round(self.resultDate[road]["generalDriveSpan"], 2))
            stopSpan.append(round(self.resultDate[road]["stopSpan"], 2))
            realStopSpan.append(round(self.resultDate[road]["realStopSpan"], 2))
            realDriveSpan.append(round(self.resultDate[road]["realDriveSpan"], 2))
            # flag = True
            for i in range(1, len(self.resultDate[road].keys())):
                k = list(self.resultDate[road].keys())[i]
                y += 1
                val = self.resultDate[road][k]
                # if flag:
                #     val = round(val)
                #     flag = False
                worksheet.write(x, y, val)
            x += 1
        workbook.close()

        # self.barDiagramDraw(path="output/" + str(factor) + "/general/"+str(factor)+"_numberOfCars.png",
        #                    title="numberOfCars", xLabel="road", yLabel="numberOfCars",
        #                    data=["numberOfCars", self.allRoads, numberOfCars])
        self.barDiagramDraw(path="output/" + str(factor) + "/general/" + str(factor) + "_generalDriveSpan.png",
                            title="generalDriveSpan", xLabel="road", yLabel="average of generalDriveSpan in S ",
                            data=["generalDriveSpan", self.allRoads, generalDriveSpan])
        self.barDiagramDraw(path="output/" + str(factor) + "/general/" + str(factor) + "_stopSpan.png",
                            title="stopSpan", xLabel="road", yLabel="average of stopSpan in S",
                            data=["stopSpan", self.allRoads, stopSpan])
        self.barDiagramDraw(path="output/" + str(factor) + "/general/" + str(factor) + "_realStopSpan.png",
                            title="realStopSpan", xLabel="road", yLabel="average of realStopSpan in S",
                            data=["realStopSpan", self.allRoads, realStopSpan])
        self.barDiagramDraw(path="output/" + str(factor) + "/general/" + str(factor) + "_realDriveSpan.png",
                            title="realDriveSpan", xLabel="road", yLabel="average of realDriveSpan in S",
                            data=["realDriveSpan", self.allRoads, realDriveSpan])

        pass


evaluator = Evaluator()

factors = ["realTime", "halfTime", "quarterTime"]
for factor in factors:
    evaluator.dataEvaluate1(factor)
    evaluator.dataEvaluate2(factor)
